"""
센서값 통계 검색 및 파일 저장 대화상자

지정된 기간의 센서 데이터 통계를 조회하고 CSV/Excel 파일로 저장합니다.
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import os
from datetime import datetime, timedelta

from ..utils.helpers import get_base_dir


class SensorStatisticsDialog:
    """센서값 통계 검색 및 파일 저장 대화상자"""

    def __init__(self, parent, app):
        """
        Args:
            parent: 부모 윈도우
            app: 메인 앱 인스턴스
        """
        self.parent = parent
        self.app = app
        self.dialog = None

        # 센서 정보 (아이콘 포함)
        self.sensor_names = {
            "co2": "🏭 이산화탄소 (ppm)",
            "h2s": "☠️ 황화수소 (ppm)",
            "co": "🔥 일산화탄소 (ppm)",
            "o2": "💨 산소 (%)",
            "lel": "⚡ 가연성가스 (%)",
            "smoke": "🌫️ 연기 (ppm)",
            "temperature": "🌡️ 온도 (℃)",
            "humidity": "💧 습도 (%)",
            "water": "🚿 누수 감지"
        }

        # 입력 변수
        self.start_date_var = None
        self.end_date_var = None
        self.interval_var = None
        self.sensor_vars = {}
        self.panel_var = None

        # 결과 데이터
        self.result_data = []

    def show(self):
        """대화상자 표시"""
        self.dialog = tk.Toplevel(self.parent)
        self.dialog.title("센서값 통계 검색")
        self.dialog.geometry("1050x720")  # 폭 확장, 높이 축소 (센서 선택 2줄로 변경)
        self.dialog.configure(bg="#F5F5F5")
        self.dialog.transient(self.parent)
        self.dialog.grab_set()

        # 중앙 배치
        self.dialog.update_idletasks()
        x = (self.dialog.winfo_screenwidth() // 2) - (525)
        y = (self.dialog.winfo_screenheight() // 2) - (360)
        self.dialog.geometry(f"1050x720+{x}+{y}")

        self._create_ui()

    def _create_ui(self):
        """UI 생성"""
        # 제목
        title_frame = tk.Frame(self.dialog, bg="#2C3E50", height=60)
        title_frame.pack(fill="x")
        title_frame.pack_propagate(False)

        tk.Label(title_frame, text="센서값 통계 검색",
                font=("Pretendard", 18, "bold"), bg="#2C3E50", fg="#FFFFFF").pack(pady=15)

        # 검색 조건 프레임
        search_frame = ttk.LabelFrame(self.dialog, text="검색 조건", padding=15)
        search_frame.pack(fill="x", padx=20, pady=15)

        # 패널 선택
        panel_frame = ttk.Frame(search_frame)
        panel_frame.pack(fill="x", pady=5)

        ttk.Label(panel_frame, text="센서 패널:", font=("Pretendard", 11)).pack(side="left")
        self.panel_var = tk.StringVar()
        self.panel_combo = ttk.Combobox(panel_frame, textvariable=self.panel_var,
                                        state="readonly", width=40, font=("Pretendard", 11))
        self.panel_combo.pack(side="left", padx=10)
        self._load_panels()

        # 기간 선택
        date_frame = ttk.Frame(search_frame)
        date_frame.pack(fill="x", pady=10)

        ttk.Label(date_frame, text="검색 기간:", font=("Pretendard", 11)).pack(side="left")

        # 시작일
        self.start_date_var = tk.StringVar(value=(datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d"))
        ttk.Entry(date_frame, textvariable=self.start_date_var, width=12,
                 font=("Pretendard", 11)).pack(side="left", padx=5)

        ttk.Label(date_frame, text="~", font=("Pretendard", 11)).pack(side="left", padx=5)

        # 종료일
        self.end_date_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        ttk.Entry(date_frame, textvariable=self.end_date_var, width=12,
                 font=("Pretendard", 11)).pack(side="left", padx=5)

        # 빠른 선택 버튼
        quick_frame = ttk.Frame(date_frame)
        quick_frame.pack(side="left", padx=20)

        ttk.Button(quick_frame, text="오늘", width=6,
                  command=lambda: self._set_quick_date(0)).pack(side="left", padx=2)
        ttk.Button(quick_frame, text="1주일", width=6,
                  command=lambda: self._set_quick_date(7)).pack(side="left", padx=2)
        ttk.Button(quick_frame, text="1개월", width=6,
                  command=lambda: self._set_quick_date(30)).pack(side="left", padx=2)
        ttk.Button(quick_frame, text="3개월", width=6,
                  command=lambda: self._set_quick_date(90)).pack(side="left", padx=2)
        ttk.Button(quick_frame, text="1년", width=6,
                  command=lambda: self._set_quick_date(365)).pack(side="left", padx=2)

        # 검색 간격 선택
        interval_frame = ttk.Frame(search_frame)
        interval_frame.pack(fill="x", pady=10)

        ttk.Label(interval_frame, text="검색 간격:", font=("Pretendard", 11)).pack(side="left")
        self.interval_var = tk.StringVar(value="1분")
        interval_combo = ttk.Combobox(interval_frame, textvariable=self.interval_var,
                                      state="readonly", width=10, font=("Pretendard", 11),
                                      values=["1분", "10분", "1시간"])
        interval_combo.pack(side="left", padx=10)

        ttk.Label(interval_frame, text="(검색 결과 그룹화 단위)", font=("Pretendard", 9),
                 foreground="#666666").pack(side="left", padx=5)

        # 센서 선택
        sensor_frame = ttk.LabelFrame(search_frame, text="센서 항목 선택", padding=10)
        sensor_frame.pack(fill="x", pady=10)

        sensor_grid = ttk.Frame(sensor_frame)
        sensor_grid.pack(fill="x")

        for i, (key, name) in enumerate(self.sensor_names.items()):
            row = i // 5  # 한 줄에 5개씩 (2줄로 표시)
            col = i % 5

            var = tk.BooleanVar(value=True)
            check = ttk.Checkbutton(sensor_grid, text=name, variable=var)
            check.grid(row=row, column=col, sticky="w", padx=8, pady=2)
            self.sensor_vars[key] = var

        # 전체 선택/해제 버튼
        btn_frame = ttk.Frame(sensor_frame)
        btn_frame.pack(fill="x", pady=(10, 0))

        ttk.Button(btn_frame, text="전체 선택", width=10,
                  command=self._select_all_sensors).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="전체 해제", width=10,
                  command=self._deselect_all_sensors).pack(side="left", padx=5)

        # 검색 버튼
        search_btn_frame = ttk.Frame(self.dialog)
        search_btn_frame.pack(fill="x", padx=20, pady=10)

        tk.Button(search_btn_frame, text="검색", command=self._search,
                 bg="#3498DB", fg="#FFFFFF", font=("Pretendard", 12, "bold"),
                 relief="raised", bd=3, width=15, height=2,
                 activebackground="#2980B9", activeforeground="#FFFFFF").pack(side="left", padx=5)

        tk.Button(search_btn_frame, text="파일로 저장", command=self._save_to_file,
                 bg="#27AE60", fg="#FFFFFF", font=("Pretendard", 12, "bold"),
                 relief="raised", bd=3, width=15, height=2,
                 activebackground="#229954", activeforeground="#FFFFFF").pack(side="left", padx=5)

        tk.Button(search_btn_frame, text="닫기", command=self.dialog.destroy,
                 bg="#95A5A6", fg="#FFFFFF", font=("Pretendard", 12, "bold"),
                 relief="raised", bd=3, width=10, height=2,
                 activebackground="#7F8C8D", activeforeground="#FFFFFF").pack(side="right", padx=5)

        # 결과 표시 영역
        result_frame = ttk.LabelFrame(self.dialog, text="검색 결과", padding=10)
        result_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))

        # 결과 테이블 (검색 결과가 최대한 많이 보이도록 height 증가)
        columns = ("sensor", "interval", "min", "max", "avg", "count", "start", "end")
        self.result_tree = ttk.Treeview(result_frame, columns=columns, show="headings", height=15)

        self.result_tree.heading("sensor", text="센서 항목")
        self.result_tree.heading("interval", text="검색간격")
        self.result_tree.heading("min", text="최소값")
        self.result_tree.heading("max", text="최대값")
        self.result_tree.heading("avg", text="평균값")
        self.result_tree.heading("count", text="데이터 수")
        self.result_tree.heading("start", text="시작일")
        self.result_tree.heading("end", text="종료일")

        self.result_tree.column("sensor", width=130, anchor="w")
        self.result_tree.column("interval", width=70, anchor="center")
        self.result_tree.column("min", width=90, anchor="center")
        self.result_tree.column("max", width=90, anchor="center")
        self.result_tree.column("avg", width=90, anchor="center")
        self.result_tree.column("count", width=70, anchor="center")
        self.result_tree.column("start", width=100, anchor="center")
        self.result_tree.column("end", width=100, anchor="center")

        # 스크롤바
        scrollbar = ttk.Scrollbar(result_frame, orient="vertical", command=self.result_tree.yview)
        self.result_tree.configure(yscrollcommand=scrollbar.set)

        self.result_tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

    def _load_panels(self):
        """연결된 패널 목록 로드"""
        panels = []
        for panel_key in self.app.panels.keys():
            if panel_key != "__waiting__":
                panels.append(panel_key)

        if panels:
            self.panel_combo['values'] = panels
            self.panel_combo.current(0)
        else:
            self.panel_combo['values'] = ["연결된 센서 없음"]

    def _set_quick_date(self, days):
        """빠른 날짜 설정"""
        end_date = datetime.now()
        start_date = end_date - timedelta(days=days)
        self.start_date_var.set(start_date.strftime("%Y-%m-%d"))
        self.end_date_var.set(end_date.strftime("%Y-%m-%d"))

    def _select_all_sensors(self):
        """모든 센서 선택"""
        for var in self.sensor_vars.values():
            var.set(True)

    def _deselect_all_sensors(self):
        """모든 센서 해제"""
        for var in self.sensor_vars.values():
            var.set(False)

    def _search(self):
        """검색 실행"""
        # 입력값 검증
        panel_key = self.panel_var.get()
        if not panel_key or panel_key == "연결된 센서 없음":
            messagebox.showwarning("알림", "센서 패널을 선택하세요.", parent=self.dialog)
            return

        try:
            start_date = datetime.strptime(self.start_date_var.get(), "%Y-%m-%d")
            end_date = datetime.strptime(self.end_date_var.get(), "%Y-%m-%d")
        except ValueError:
            messagebox.showerror("오류", "날짜 형식이 올바르지 않습니다.\nYYYY-MM-DD 형식으로 입력하세요.", parent=self.dialog)
            return

        if start_date > end_date:
            messagebox.showerror("오류", "시작일이 종료일보다 늦을 수 없습니다.", parent=self.dialog)
            return

        # 검색 간격
        interval_text = self.interval_var.get()
        interval_minutes = {"1분": 1, "10분": 10, "1시간": 60}.get(interval_text, 1)

        # 선택된 센서
        selected_sensors = [key for key, var in self.sensor_vars.items() if var.get()]
        if not selected_sensors:
            messagebox.showwarning("알림", "검색할 센서 항목을 선택하세요.", parent=self.dialog)
            return

        # SID와 peer 추출
        if "@" in panel_key:
            sid = panel_key.split("@")[0]
            peer = panel_key.split("@")[1] if "@" in panel_key else ""
        elif "#" in panel_key:
            sid = panel_key.split("#")[0]
            peer = ""
        else:
            sid = panel_key
            peer = ""

        # 검색 실행
        self.result_data = []
        self.result_tree.delete(*self.result_tree.get_children())

        try:
            for sensor_key in selected_sensors:
                stats = self._get_sensor_stats(sid, peer, sensor_key, start_date, end_date, interval_minutes)
                if stats:
                    sensor_name = self.sensor_names.get(sensor_key, sensor_key)
                    row_data = {
                        "sensor_key": sensor_key,
                        "sensor_name": sensor_name,
                        "interval": interval_text,
                        "min": stats["min"],
                        "max": stats["max"],
                        "avg": stats["avg"],
                        "count": stats["count"],
                        "start": start_date.strftime("%Y-%m-%d"),
                        "end": end_date.strftime("%Y-%m-%d")
                    }
                    self.result_data.append(row_data)

                    # 트리뷰에 추가
                    self.result_tree.insert("", "end", values=(
                        sensor_name,
                        interval_text,
                        f"{stats['min']:.2f}" if stats['min'] is not None else "-",
                        f"{stats['max']:.2f}" if stats['max'] is not None else "-",
                        f"{stats['avg']:.2f}" if stats['avg'] is not None else "-",
                        stats['count'],
                        start_date.strftime("%Y-%m-%d"),
                        end_date.strftime("%Y-%m-%d")
                    ))

            if not self.result_data:
                messagebox.showinfo("알림", "해당 기간에 데이터가 없습니다.", parent=self.dialog)

        except Exception as e:
            messagebox.showerror("오류", f"검색 중 오류가 발생했습니다:\n{e}", parent=self.dialog)

    def _get_sensor_stats(self, sid, peer, sensor_key, start_date, end_date, interval_minutes=1):
        """지정 기간의 센서 통계 조회 (간격별 샘플링)

        Args:
            interval_minutes: 샘플링 간격 (1분, 10분, 60분)
                - 1분: 모든 데이터 사용
                - 10분/60분: 해당 간격별 평균값으로 샘플링하여 통계 계산
        """
        import sqlite3
        import time

        try:
            # 로그 매니저에서 DB 경로 가져오기
            db_path = self.app.logs.db_path

            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()

            peer_ip = peer.split(":")[0] if peer else ""

            # 날짜를 타임스탬프로 변환
            start_ts = time.mktime(start_date.timetuple())
            end_ts = time.mktime((end_date + timedelta(days=1)).timetuple())

            # 유효한 값 필터링 조건
            if sensor_key == "temperature":
                filter_condition = f"AND {sensor_key} IS NOT NULL AND {sensor_key} != -1 AND {sensor_key} BETWEEN -50 AND 50"
            elif sensor_key == "water":
                filter_condition = f"AND {sensor_key} IS NOT NULL"
            else:
                filter_condition = f"AND {sensor_key} IS NOT NULL AND {sensor_key} != -1 AND {sensor_key} >= 0"

            # 간격에 따른 쿼리 (간격별 그룹화)
            interval_seconds = interval_minutes * 60

            if interval_minutes > 1:
                # 간격별로 그룹화하여 평균을 구한 후, 전체 통계 계산
                # (timestamp / interval) 단위로 그룹화
                query = f"""
                    SELECT
                        MIN(avg_val) as min_val,
                        MAX(avg_val) as max_val,
                        AVG(avg_val) as avg_val,
                        COUNT(*) as count_val
                    FROM (
                        SELECT
                            CAST(timestamp / {interval_seconds} AS INTEGER) as time_bucket,
                            AVG({sensor_key}) as avg_val
                        FROM sensor_data
                        WHERE sid = ? AND peer_ip = ?
                        AND timestamp >= ? AND timestamp < ?
                        {filter_condition}
                        GROUP BY time_bucket
                    )
                """
            else:
                # 1분 간격: 모든 데이터 사용
                query = f"""
                    SELECT
                        MIN({sensor_key}) as min_val,
                        MAX({sensor_key}) as max_val,
                        AVG({sensor_key}) as avg_val,
                        COUNT({sensor_key}) as count_val
                    FROM sensor_data
                    WHERE sid = ? AND peer_ip = ?
                    AND timestamp >= ? AND timestamp < ?
                    {filter_condition}
                """

            cursor.execute(query, (sid, peer_ip, start_ts, end_ts))
            row = cursor.fetchone()
            conn.close()

            if row and row[3] > 0:
                return {
                    "min": row[0],
                    "max": row[1],
                    "avg": row[2],
                    "count": row[3]
                }

        except Exception as e:
            print(f"통계 조회 오류: {e}")

        return None

    def _save_to_file(self):
        """결과를 파일로 저장"""
        if not self.result_data:
            messagebox.showwarning("알림", "먼저 검색을 실행하세요.", parent=self.dialog)
            return

        # 기본 저장 디렉토리: 설치 경로/statistics
        install_dir = get_base_dir()
        statistics_dir = os.path.join(install_dir, "statistics")
        if not os.path.exists(statistics_dir):
            os.makedirs(statistics_dir, exist_ok=True)

        # 파일 형식 선택
        self.dialog.attributes("-topmost", True)
        self.dialog.update()

        filepath = filedialog.asksaveasfilename(
            title="통계 결과 저장",
            defaultextension=".xlsx",
            filetypes=[
                ("Excel 파일 (원시데이터+그래프)", "*.xlsx"),
                ("CSV 파일 (요약만)", "*.csv"),
                ("모든 파일", "*.*")
            ],
            initialfile=f"sensor_stats_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
            initialdir=statistics_dir,
            parent=self.dialog
        )

        self.dialog.attributes("-topmost", False)

        if not filepath:
            return

        try:
            if filepath.endswith(".xlsx"):
                self._save_as_excel(filepath)
            else:
                self._save_as_csv(filepath)

            # 완료 메시지
            self._show_save_complete_dialog(filepath)

        except Exception as e:
            messagebox.showerror("오류", f"파일 저장 중 오류가 발생했습니다:\n{e}", parent=self.dialog)

    def _save_as_csv(self, filepath):
        """CSV 파일로 저장"""
        import csv

        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)

            # 헤더
            writer.writerow(["센서 항목", "검색간격", "최소값", "최대값", "평균값", "데이터 수", "시작일", "종료일"])

            # 데이터
            for row in self.result_data:
                writer.writerow([
                    row["sensor_name"],
                    row.get("interval", "1분"),
                    f"{row['min']:.2f}" if row['min'] is not None else "",
                    f"{row['max']:.2f}" if row['max'] is not None else "",
                    f"{row['avg']:.2f}" if row['avg'] is not None else "",
                    row['count'],
                    row['start'],
                    row['end']
                ])

    def _get_raw_sensor_data(self, sid, peer, sensor_key, start_date, end_date, interval_minutes=1):
        """원시 센서 데이터 조회 (그래프 및 탭별 저장용)"""
        import sqlite3
        import time

        try:
            db_path = self.app.logs.db_path
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()

            peer_ip = peer.split(":")[0] if peer else ""
            start_ts = time.mktime(start_date.timetuple())
            end_ts = time.mktime((end_date + timedelta(days=1)).timetuple())

            # 유효한 값 필터링 조건
            if sensor_key == "temperature":
                filter_condition = f"AND {sensor_key} IS NOT NULL AND {sensor_key} != -1 AND {sensor_key} BETWEEN -50 AND 50"
            elif sensor_key == "water":
                filter_condition = f"AND {sensor_key} IS NOT NULL"
            else:
                filter_condition = f"AND {sensor_key} IS NOT NULL AND {sensor_key} != -1 AND {sensor_key} >= 0"

            interval_seconds = interval_minutes * 60

            if interval_minutes > 1:
                # 간격별 평균
                query = f"""
                    SELECT
                        CAST(timestamp / {interval_seconds} AS INTEGER) * {interval_seconds} as time_bucket,
                        AVG({sensor_key}) as value
                    FROM sensor_data
                    WHERE sid = ? AND peer_ip = ?
                    AND timestamp >= ? AND timestamp < ?
                    {filter_condition}
                    GROUP BY CAST(timestamp / {interval_seconds} AS INTEGER)
                    ORDER BY time_bucket
                """
            else:
                # 원시 데이터
                query = f"""
                    SELECT timestamp, {sensor_key}
                    FROM sensor_data
                    WHERE sid = ? AND peer_ip = ?
                    AND timestamp >= ? AND timestamp < ?
                    {filter_condition}
                    ORDER BY timestamp
                """

            cursor.execute(query, (sid, peer_ip, start_ts, end_ts))
            rows = cursor.fetchall()
            conn.close()

            return rows

        except Exception as e:
            print(f"원시 데이터 조회 오류: {e}")
            return []

    def _save_as_excel(self, filepath):
        """Excel 파일로 저장 (검색결과 + 센서별 원시데이터 탭 + 그래프)"""
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.chart import LineChart, Reference
            from openpyxl.utils.dataframe import dataframe_to_rows
            from io import BytesIO
            import time

            # 검색 조건 파싱
            panel_key = self.panel_var.get()
            if "@" in panel_key:
                sid = panel_key.split("@")[0]
                peer = panel_key.split("@")[1]
            elif "#" in panel_key:
                sid = panel_key.split("#")[0]
                peer = ""
            else:
                sid = panel_key
                peer = ""

            start_date = datetime.strptime(self.start_date_var.get(), "%Y-%m-%d")
            end_date = datetime.strptime(self.end_date_var.get(), "%Y-%m-%d")
            interval_text = self.interval_var.get()
            interval_minutes = {"1분": 1, "10분": 10, "1시간": 60}.get(interval_text, 1)

            # Workbook 생성
            wb = Workbook()

            # ====== 첫 번째 시트: 검색결과 요약 ======
            ws_summary = wb.active
            ws_summary.title = "검색결과"

            # 검색 조건 정보
            ws_summary.append(["센서값 통계 검색 결과"])
            ws_summary.append([])
            ws_summary.append(["검색 조건"])
            ws_summary.append(["센서 패널:", panel_key])
            ws_summary.append(["검색 기간:", f"{self.start_date_var.get()} ~ {self.end_date_var.get()}"])
            ws_summary.append(["검색 간격:", interval_text])
            ws_summary.append([])

            # 결과 테이블 헤더
            ws_summary.append(["센서 항목", "검색간격", "최소값", "최대값", "평균값", "데이터 수", "시작일", "종료일"])

            # 결과 데이터
            for row in self.result_data:
                ws_summary.append([
                    row["sensor_name"],
                    row.get("interval", "1분"),
                    round(row["min"], 2) if row["min"] is not None else None,
                    round(row["max"], 2) if row["max"] is not None else None,
                    round(row["avg"], 2) if row["avg"] is not None else None,
                    row["count"],
                    row["start"],
                    row["end"]
                ])

            # ====== 각 센서별 원시데이터 시트 생성 ======
            sensor_data_sheets = {}
            selected_sensors = [key for key, var in self.sensor_vars.items() if var.get()]

            for sensor_key in selected_sensors:
                sensor_name = self.sensor_names.get(sensor_key, sensor_key)
                raw_data = self._get_raw_sensor_data(sid, peer, sensor_key, start_date, end_date, interval_minutes)

                if raw_data:
                    # 시트 이름 생성 (Excel 시트 이름 제한: 31자, 특수문자 제외)
                    sheet_name = sensor_name[:20].replace("/", "_").replace("(", "").replace(")", "").replace(" ", "_")
                    ws = wb.create_sheet(title=sheet_name)

                    # 헤더
                    ws.append(["시간", sensor_name])

                    # 데이터
                    for ts, value in raw_data:
                        dt_str = datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S")
                        ws.append([dt_str, round(value, 2) if value is not None else None])

                    sensor_data_sheets[sensor_key] = {
                        "sheet": ws,
                        "name": sensor_name,
                        "data_count": len(raw_data)
                    }

            # ====== 전체 센서 그래프 시트 ======
            ws_all_graph = wb.create_sheet(title="전체센서_그래프")
            ws_all_graph.append(["전체 센서 데이터 그래프"])
            ws_all_graph.append([])

            # 전체 그래프용 데이터 수집 (시간 기준 정렬)
            all_times = set()
            sensor_values = {}

            for sensor_key in selected_sensors:
                raw_data = self._get_raw_sensor_data(sid, peer, sensor_key, start_date, end_date, interval_minutes)
                sensor_values[sensor_key] = {ts: val for ts, val in raw_data}
                all_times.update([ts for ts, _ in raw_data])

            if all_times:
                sorted_times = sorted(all_times)

                # 헤더 (센서명 사용)
                header = ["시간"] + [self.sensor_names.get(k, k) for k in selected_sensors]
                ws_all_graph.append(header)

                # 데이터 (최대 1000행으로 제한하여 파일 크기 관리)
                max_rows = min(len(sorted_times), 1000)
                step = max(1, len(sorted_times) // max_rows)

                data_start_row = 4  # 데이터 시작 행 (헤더 포함)
                data_rows = 0

                for i in range(0, len(sorted_times), step):
                    if data_rows >= max_rows:
                        break
                    ts = sorted_times[i]
                    dt_str = datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M")
                    row_values = [dt_str]
                    for sensor_key in selected_sensors:
                        val = sensor_values.get(sensor_key, {}).get(ts)
                        row_values.append(round(val, 2) if val is not None else None)
                    ws_all_graph.append(row_values)
                    data_rows += 1

                # 전체 그래프 생성 (이중 Y축: CO2는 우측, 나머지는 좌측 0-100)
                if data_rows > 1:
                    from openpyxl.chart.series import SeriesLabel
                    from openpyxl.drawing.line import LineProperties
                    from openpyxl.chart.shapes import GraphicalProperties

                    # 센서별 색상 정의
                    sensor_colors = {
                        "co2": "1E88E5",      # 파랑
                        "h2s": "D81B60",      # 핑크
                        "co": "FFC107",       # 노랑
                        "o2": "43A047",       # 녹색
                        "lel": "FF5722",      # 주황
                        "smoke": "6A1B9A",    # 보라
                        "temperature": "E53935",  # 빨강
                        "humidity": "00ACC1",    # 청록
                        "water": "8D6E63"     # 갈색
                    }

                    # CO2와 기타 센서 분리
                    co2_sensors = [k for k in selected_sensors if k == "co2"]
                    other_sensors = [k for k in selected_sensors if k != "co2"]

                    # 기본 차트 (기타 센서용 - 좌측 Y축, 0-100 스케일)
                    chart1 = LineChart()
                    chart1.title = "전체 센서 추이 (CO2: 우측축 / 기타: 좌측축 0-100)"
                    chart1.style = 10
                    chart1.x_axis.title = "시간"
                    chart1.y_axis.title = "기타 센서값 (0-100)"
                    chart1.y_axis.scaling.min = 0
                    chart1.y_axis.scaling.max = 100
                    chart1.width = 25
                    chart1.height = 14

                    # 기타 센서 데이터 추가 (좌측 Y축)
                    for sensor_key in other_sensors:
                        col_idx = selected_sensors.index(sensor_key) + 2
                        data = Reference(ws_all_graph, min_col=col_idx, min_row=data_start_row - 1,
                                        max_col=col_idx, max_row=data_start_row + data_rows - 1)
                        chart1.add_data(data, titles_from_data=True)

                        series = chart1.series[-1]
                        color = sensor_colors.get(sensor_key, "000000")
                        series.graphicalProperties.line.solidFill = color
                        series.graphicalProperties.line.width = 25000

                    # 카테고리 (시간)
                    cats = Reference(ws_all_graph, min_col=1, min_row=data_start_row,
                                    max_row=data_start_row + data_rows - 1)
                    chart1.set_categories(cats)

                    # CO2가 있으면 보조 Y축으로 추가
                    if co2_sensors:
                        # CO2 차트 (우측 Y축)
                        chart2 = LineChart()
                        chart2.y_axis.axId = 200  # 보조 축 ID
                        chart2.y_axis.title = "CO2 (ppm)"

                        for sensor_key in co2_sensors:
                            col_idx = selected_sensors.index(sensor_key) + 2
                            data = Reference(ws_all_graph, min_col=col_idx, min_row=data_start_row - 1,
                                            max_col=col_idx, max_row=data_start_row + data_rows - 1)
                            chart2.add_data(data, titles_from_data=True)

                            series = chart2.series[-1]
                            color = sensor_colors.get(sensor_key, "1E88E5")
                            series.graphicalProperties.line.solidFill = color
                            series.graphicalProperties.line.width = 35000  # CO2는 더 굵게

                        # 보조 Y축을 우측에 배치
                        chart2.y_axis.crosses = "max"

                        # 차트 결합
                        chart1 += chart2

                    # 범례 표시
                    chart1.legend.position = 'b'

                    ws_all_graph.add_chart(chart1, "A" + str(data_start_row + data_rows + 3))

            # ====== 센서별 개별 그래프 시트 ======
            for sensor_key, sheet_info in sensor_data_sheets.items():
                ws = sheet_info["sheet"]
                sensor_name = sheet_info["name"]
                data_count = sheet_info["data_count"]

                if data_count > 1:
                    # 개별 그래프 생성
                    chart = LineChart()
                    chart.title = f"{sensor_name} 추이"
                    chart.style = 10
                    chart.x_axis.title = "시간"
                    chart.y_axis.title = sensor_name
                    chart.width = 18
                    chart.height = 10

                    # 데이터 범위 (최대 500행)
                    max_chart_rows = min(data_count, 500)

                    data = Reference(ws, min_col=2, min_row=1, max_row=max_chart_rows + 1)
                    chart.add_data(data, titles_from_data=True)

                    cats = Reference(ws, min_col=1, min_row=2, max_row=max_chart_rows + 1)
                    chart.set_categories(cats)

                    # 그래프 위치 (데이터 옆)
                    ws.add_chart(chart, "D2")

            # 파일 저장
            wb.save(filepath)

        except ImportError as e:
            messagebox.showerror(
                "오류",
                "Excel 저장을 위해 pandas와 openpyxl 패키지가 필요합니다.\n\n"
                f"설치 명령어:\npip install pandas openpyxl\n\n오류: {e}",
                parent=self.dialog
            )
            raise

    def _show_save_complete_dialog(self, filepath):
        """저장 완료 다이얼로그"""
        folder_path = os.path.dirname(filepath)
        filename = os.path.basename(filepath)

        # 다이얼로그 생성
        complete_dialog = tk.Toplevel(self.dialog)
        complete_dialog.title("저장 완료")
        complete_dialog.configure(bg="#2C3E50")
        complete_dialog.transient(self.dialog)
        complete_dialog.grab_set()

        # 창 크기 및 중앙 배치
        dialog_width = 500
        dialog_height = 250
        complete_dialog.geometry(f"{dialog_width}x{dialog_height}")
        complete_dialog.update_idletasks()
        x = (complete_dialog.winfo_screenwidth() // 2) - (dialog_width // 2)
        y = (complete_dialog.winfo_screenheight() // 2) - (dialog_height // 2)
        complete_dialog.geometry(f"{dialog_width}x{dialog_height}+{x}+{y}")

        complete_dialog.attributes("-topmost", True)
        complete_dialog.focus_force()

        # 제목
        tk.Label(complete_dialog, text="저장 완료",
                font=("Pretendard", 18, "bold"), bg="#2C3E50", fg="#27AE60").pack(pady=20)

        # 정보 프레임
        info_frame = tk.Frame(complete_dialog, bg="#34495E", relief="raised", bd=2)
        info_frame.pack(fill="x", padx=30, pady=10)

        # 저장 디렉토리
        dir_frame = tk.Frame(info_frame, bg="#34495E")
        dir_frame.pack(fill="x", padx=15, pady=(15, 5))

        tk.Label(dir_frame, text="저장 위치:", font=("Pretendard", 10, "bold"),
                bg="#34495E", fg="#FFD700").pack(side="left")
        tk.Label(dir_frame, text=folder_path, font=("Pretendard", 9),
                bg="#34495E", fg="#FFFFFF", wraplength=350, justify="left").pack(side="left", padx=10)

        # 파일명
        file_frame = tk.Frame(info_frame, bg="#34495E")
        file_frame.pack(fill="x", padx=15, pady=(5, 15))

        tk.Label(file_frame, text="파일명:", font=("Pretendard", 10, "bold"),
                bg="#34495E", fg="#FFD700").pack(side="left")
        tk.Label(file_frame, text=filename, font=("Pretendard", 9),
                bg="#34495E", fg="#FFFFFF").pack(side="left", padx=10)

        # 버튼
        button_frame = tk.Frame(complete_dialog, bg="#2C3E50")
        button_frame.pack(pady=15)

        tk.Button(button_frame, text="확인", command=complete_dialog.destroy,
                 bg="#27AE60", fg="#FFFFFF", font=("Pretendard", 11, "bold"),
                 relief="raised", bd=3, width=10,
                 activebackground="#229954", activeforeground="#FFFFFF").pack()

        complete_dialog.bind("<Return>", lambda e: complete_dialog.destroy())
        complete_dialog.bind("<Escape>", lambda e: complete_dialog.destroy())
